import datetime 
import locale
import openpyxl

#!ДАТА ПЕРЕДЕАТСЯ ЗА КЛАССОМ , хранится как string в виде '%d.%m.%Y'

class Teacher():
    #на вход группа Института ИТ 
    def __init__(self,teacher):
        self.teacher = teacher
        self.date = None

        self.book = None
        self.sheet =  None

    def get_diaposon_learning_date(self):
        cell = self.sheet.cell(row=1, column=1)
        start_str = cell.value.split('с ')[1].split(' по ')[0] #начало обучения
        end_str = cell.value.split('по ')[1] #конец обучения 
        return start_str,end_str
    
    def is_even_week(self):
        first_date, second_date = self.get_diaposon_learning_date()
        start_date = datetime.datetime.strptime(first_date, '%d.%m.%Y').date()
        date_object = datetime.datetime.strptime(self.date, '%d.%m.%Y').date()
        start_week_day = start_date.isoweekday()
        diff_days = (date_object - start_date).days
        week_number = (diff_days + start_week_day - 1) // 7 + 1 #какая учебная неделя
        is_even = week_number % 2 == 1
        return is_even
    
    def get_day_of_week(self):
        date_object = datetime.datetime.strptime(self.date, '%d.%m.%Y')
        locale.setlocale(locale.LC_ALL, 'ru_RU.UTF-8')
        day_of_week = date_object.strftime('%A')
        return day_of_week
    
    def get_month_name_genitive(self):
        month_names_genitive = [
            "января",
            "февраля",
            "марта",
            "апреля",
            "мая",
            "июня",
            "июля",
            "августа",
            "сентября",
            "октября",
            "ноября",
            "декабря"
        ]

        date = datetime.datetime.strptime(self.date, '%d.%m.%Y')
        month_number = date.month - 1

        return f'{date.day} {month_names_genitive[month_number]}' 

    def found_teacher(self):
        #Список с строками каждый это день недели
        all_path = ["II\Oznokom praktika\BlockF\XLSX_files\IIT_1-kurs_22_23_vesna_27.04.2023.xlsx", "II\Oznokom praktika\BlockF\XLSX_files\IIT_2-kurs_22_23_vesna_15.05.2023.xlsx","II\Oznokom praktika\BlockF\XLSX_files\IIT_3-kurs_22_23_vesna_22.05.2023.xlsx"]
        formatted_date = self.get_month_name_genitive()

        list_with_days = [None,None,None,None,None,None,None]
        for path in all_path:
            self.book = openpyxl.load_workbook(path)
            self.sheet =  self.book.active 
            
            #проход по 3 ряду
            row = 3
            cells = self.sheet[row]
            for cell in cells:
                locale.setlocale(locale.LC_ALL, 'ru_RU.UTF-8')
                day_of_the_week = self.get_day_of_week()
                base_row = cell.row
                base_column = cell.column
                #определение начального положения положения клетки(зависит от дня) 
                days_of_week = {
                    "понедельник": 0,
                    "вторник": 1,
                    "среда": 2,
                    "четверг": 3,
                    "пятница": 4,
                    "суббота": 5,
                    "воскресенье": 6
                }
                value = days_of_week[day_of_the_week]

                if value == 6:
                    return f"У преподавателя {self.teacher} , на {formatted_date}"

                even = self.is_even_week()
                if even:
                    shift_even = 2
                else:
                    shift_even = 1
                    
                shift_day = 14
                #влево враво column
                #вверх-вниз row
                for i in range(0,14,2):
                    all_shift = shift_day *value + shift_even+i
                    cell_current = self.sheet.cell(column = base_column, row= all_shift+base_row)
                    if cell_current.value == self.teacher:
                        subject = cell_current.offset(column = -2).value
                        type_of_activity = cell_current.offset(column = -1).value
                        group = cell_current.offset(column = -2, row = -(all_shift+1) ).value
                        class_where = cell_current.offset(column = +1, row = 0).value
                        string = f' {subject} , {type_of_activity} , {group} , {class_where}\n'
                        if i == 0:
                            list_with_days[0] = string
                        else:
                            list_with_days[int(i/2)] = string

        string_for_output = f"Расписание преподавателя {self.teacher} на {formatted_date}: \n\n"
        for i in range(len(list_with_days)):
            if list_with_days[i] != None:
                string_for_output += f"{i+1}) {list_with_days[i]}"
            else:
                string_for_output += f"{i+1}) - \n"
                
        return string_for_output    
                        




def parse_schedule(surname_prepod):
    all_paths = [
        r"II\Oznokom praktika\BlockF\XLSX_files\IIT_1-kurs_22_23_vesna_27.04.2023.xlsx",
        r"II\Oznokom praktika\BlockF\XLSX_files\IIT_2-kurs_22_23_vesna_15.05.2023.xlsx",
        r"II\Oznokom praktika\BlockF\XLSX_files\IIT_3-kurs_22_23_vesna_22.05.2023.xlsx"
    ]
    all_teachers = []
    teacher_found = False

    for document_path in all_paths:
        workbook = openpyxl.load_workbook(document_path)
        sheet = workbook.active

        teachers = []

        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if cell is None:
                    continue

                split_cell = cell.split()
                if len(split_cell) > 0 and surname_prepod == split_cell[0]:
                    teacher_name = cell
                    teachers.append(teacher_name)
                    teacher_found = True

        unique_teachers = list(set(teachers))
        all_teachers.extend(unique_teachers)

    if teacher_found:
        return list(set(all_teachers))
    else:
        return None
                    