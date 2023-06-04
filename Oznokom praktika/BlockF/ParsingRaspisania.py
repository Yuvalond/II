#только институт ИТ
import datetime 
import locale
import requests
import openpyxl
import os
import re
import json
from bs4 import BeautifulSoup

class Raspisanie():
    #на вход группа Института ИТ 
    def __init__(self, group : str , id_user):
        self.group = group                                          #Группа
        self.course = self.get_current_course()                     #Курс
        self.date = None                                            #Текущая дата , где находимся
        self.path = self.get_xlsx_from_mirea()                      #Путь к файлу для текущего курса
        self.id_user = id_user                                      #айди пользователя вконтакте , для сохранения данных и выбора

        self.book = openpyxl.load_workbook(self.path, read_only=True)
        self.sheet =  self.book.active 
    

    def to_json(self):
        return {
            'group': self.group,
            'course': self.course,
            'date': self.date,
            'path': self.path,
            'id_user': self.id_user
        }

    #на каком курсе сейчас человек
    def get_current_course(self):
        year_of_admission = int(self.group.split('-')[-1])  # получаем год поступления из строки и преобразуем в целое число
        # определяем дату 1 сентября текущего года
        new_school_year = datetime.date(datetime.date.today().year, 9, 1)
        if datetime.date.today() < new_school_year:
            course = datetime.date.today().year - (year_of_admission + 2000)
        else:
            course = datetime.date.today().year - (year_of_admission + 2000) + 1
        return course
     

    #получаем xlsx файл и возвращаем его путь 
    def get_xlsx_from_mirea(self):
            page = requests.get("https://www.mirea.ru/schedule/")
            soup = BeautifulSoup(page.text, "html.parser")
            if self.course > 3:
                  return print(f"Курс слишкои большой ")
            result = soup.find ("div", {"class": "schedule"}).\
                    find(string = "Институт информационных технологий").\
                    find_parent("div").\
                    find_parent("div").\
                    findAll('a') #получить ссылки
            if not os.path.exists("II/Oznokom praktika/BlockF/XLSX_files"):
                os.makedirs("II/Oznokom praktika/BlockF/XLSX_files")
            for link in result:
                    link = link['href']                                     # получение самой ссылки
                    if link == 'javascript:void(0)':                        # пропускаем эту строку
                            continue
                    name_link = "IIT" + link.split("IIT")[-1]               #получения названия файла 
                    num_course = int(name_link.split("_")[1].split("-")[0]) #получение названия курса данного файла
                    pattern = r"^[\w-]+\_\d{2}\.\d{2}\.\d{4}\.xlsx$"
                    if re.match(pattern,name_link):  
                        if self.course == num_course:
                            path = f"II/Oznokom praktika/BlockF/XLSX_files/{name_link}"
                            with open (path , "wb") as f:
                                    resp = requests.get(link)
                                    f.write(resp.content)
                            return path



    #получение из файла даты обучения
    def get_diaposon_learning_date(self):
        cell = self.sheet.cell(row=1, column=1)
        start_str = cell.value.split('с ')[1].split(' по ')[0] #начало обучения
        end_str = cell.value.split('по ')[1] #конец обучения 
        return start_str,end_str

    #получаем день недели str
    def get_day_of_week(self):
        date_object = datetime.datetime.strptime(self.date, '%d.%m.%Y')
        locale.setlocale(locale.LC_ALL, 'ru_RU.UTF-8')
        day_of_week = date_object.strftime('%A')
        return day_of_week

    #четная неделя или нет + там есть порядковый номер недели
    def is_even_week(self):
        first_date, second_date = self.get_diaposon_learning_date()
        start_date = datetime.datetime.strptime(first_date, '%d.%m.%Y').date()
        end_date = datetime.datetime.strptime(second_date, '%d.%m.%Y').date()
        date_object = datetime.datetime.strptime(self.date, '%d.%m.%Y').date()
        if date_object < start_date or date_object > end_date:
            return "Мы не учимся брат"
        start_week_day = start_date.isoweekday()
        diff_days = (date_object - start_date).days
        week_number = (diff_days + start_week_day - 1) // 7 + 1 #какая учебная неделя
        is_even = week_number % 2 == 0
        return is_even
    
    def get_number_of_week(self):
        first_date, second_date = self.get_diaposon_learning_date()
        start_date = datetime.datetime.strptime(first_date, '%d.%m.%Y').date()
        today = datetime.date.today()
        start_week_day = start_date.isoweekday()
        diff_days = (today - start_date).days
        week_number = (diff_days + start_week_day - 1) // 7 + 1
        return f'Идет {week_number} неделя.'

    #получение ячейки группы в файле 
    def find_group_in_file(self):
        row = self.sheet[2] 
        for cell in row:
            if cell.value == self.group:
                return cell
        return None

    #Вывод в родительном падеже
    def get_month_name_genitive(self):
        month_names_genitive = [
            "января",
            "февраля",
            "марта",
            "апреля",
            "мая",
            "июня",
            "июля",
            "августа",
            "сентября",
            "октября",
            "ноября",
            "декабря"
        ]

        date = datetime.datetime.strptime(self.date, '%d.%m.%Y')
        month_number = date.month - 1

        return f'{date.day} {month_names_genitive[month_number]}' 

    #получение листа с расписанем на день расписания 
    def get_raspisanie(self, cell): # путь к файлу xml, ячейка, дата
        locale.setlocale(locale.LC_ALL, 'ru_RU.UTF-8')
        base_row = cell.row
        base_column = cell.column
        uneven = self.is_even_week()
        day_of_the_week = self.get_day_of_week()
        #определение начального положения положения клетки(зависит от дня) 
        days_of_week = {
            "понедельник": 0,
            "вторник": 1,
            "среда": 2,
            "четверг": 3,
            "пятница": 4,
            "суббота": 5,
            "воскресенье": 6
        }
        #вывод расписания для этого дня
        formatted_date = self.get_month_name_genitive()
        date_for_list = f"Расписание на {formatted_date}:"
        value = days_of_week[day_of_the_week]
        if value == 6:
            return f"{date_for_list} \nПар нет, Чилим братья!"
        result_list = []
        result_list.append(date_for_list)
        shift_uneven = 0
        if uneven == False:
            pass
        else:
            shift_uneven += 1
        for i in range(0, 14, 2):
            row = [(self.sheet.cell(row=base_row + 14 * value + shift_uneven + i + 2, column=base_column + j).value or '-').replace("\n", " ")
                for j in range(4)]
            if row[0] == "-":
                result_list.append(f"{int(i/2 + 1)}) -")
            else:
                result_list.append(f"{int(i/2 + 1)}) {row[0]}, {row[1]}, {row[2]}, {row[3]}")
        result = '\n'.join(result_list)
        return result
    
    def get_raspisanie_for_a_week(self,cell):
        locale.setlocale(locale.LC_ALL, 'ru_RU.UTF-8')
        start_of_week = self.date - datetime.timedelta(days=self.date.weekday()) # Находим начало недели
        days_of_week = [start_of_week + datetime.timedelta(days=i) for i in range(6)]  # Создаем список из дней недели
        string  = ''
        for day in days_of_week:
            self.date = day.strftime("%d.%m.%Y")
            base_row = cell.row
            base_column = cell.column
            uneven = self.is_even_week()
            day_of_the_week = self.get_day_of_week()
            #определение начального положения положения клетки(зависит от дня) 
            days_of_week = {
                "понедельник": 0,
                "вторник": 1,
                "среда": 2,
                "четверг": 3,
                "пятница": 4,
                "суббота": 5,
                "воскресенье": 6
            }
            #вывод расписания для этого дня
            value = days_of_week[day_of_the_week]
            days_list = list(days_of_week.keys())
            day_name = days_list[value]
            formatted_date = self.get_month_name_genitive()
            date_for_list = f"Расписание на {day_name} {formatted_date}:"
            if value == 6:
                return f"{date_for_list} \nПар нет, Чилим братья!"
            result_list = []
            result_list.append(date_for_list)
            shift_uneven = 0
            if uneven == False:
                pass
            else:
                shift_uneven += 1
            for i in range(0, 14, 2):
                row = [(self.sheet.cell(row=base_row + 14 * value + shift_uneven + i + 2, column=base_column + j).value or '-').replace("\n", " ")
                    for j in range(4)]
                if row[0] == "-":
                    result_list.append(f"{int(i/2 + 1)}) —")
                else:
                    result_list.append(f"{int(i/2 + 1)}) {row[0]}, {row[1]}, {row[2]}, {row[3]}")
            result = '\n'.join(result_list)
            string += f'{result}\n\n'
        return string

    #Возврат строки распиния.
    def get_raspisanie_moment(self):
        cell = self.find_group_in_file()
        if cell is not None:
            result = self.get_raspisanie(cell)
            return result
        else:
            return f"Группа {self.group} не найдена в файле"
    
    def get_raspisanie_moment_for_week(self):
        cell = self.find_group_in_file()
        if cell is not None:
            result = self.get_raspisanie_for_a_week(cell)
            return result
        else:
            return f"Группа {self.group} не найдена в файле"
    
    def get_raspisanie_moment_for_day_of_week(self,day_of_week):
        #получить дату по дню недели
        #получить четность нашей недели:

        #если четная
            #вывести расписание на этот день + расписание через 7 дней
        #если не четная
            #вывести расписание 7 дней назад + расписание на этот день
        self.date = self.get_date_for_week_of_date(day_of_week).strftime("%d.%m.%Y")
        not_even = self.is_even_week()
        one_week = datetime.timedelta(days=7)
        string = ''
        if not_even:
            self.date = (datetime.datetime.strptime(self.date,"%d.%m.%Y") - one_week).strftime("%d.%m.%Y")
            string_func = f'{self.get_raspisanie_moment()}\n\n'
            string += string_func.replace(string_func.split('\n', 1)[0], f"Расписание на четный(ую) {day_of_week}")
            self.date = (datetime.datetime.strptime(self.date,"%d.%m.%Y") + one_week).strftime("%d.%m.%Y")
            string_func = f'{self.get_raspisanie_moment()}\n\n'
            string += string_func.replace(string_func.split('\n', 1)[0], f"Расписание на не четный(ую) {day_of_week}")
        else:
            string_func = f'{self.get_raspisanie_moment()}\n\n'
            string += string_func.replace(string_func.split('\n', 1)[0], f"Расписание на четный(ую) {day_of_week}")
            self.date = (datetime.datetime.strptime(self.date,"%d.%m.%Y") + one_week).strftime("%d.%m.%Y")
            string_func = f'{self.get_raspisanie_moment()}\n\n'
            string += string_func.replace(string_func.split('\n', 1)[0], f"Расписание на не четный(ую) {day_of_week}")
            
        return string


    
    def get_date_for_week_of_date(self,day_of_week):
        today_date = datetime.date.today()
        days = {'понедельник': 0, 'вторник': 1, 'среда': 2, 'четверг': 3, 'пятница': 4, 'суббота': 5}
        days_diff = days.get(day_of_week) - today_date.weekday()
        result_date = today_date + datetime.timedelta(days=days_diff)
        return result_date