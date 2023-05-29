

# # files = os.listdir('/path/to/directory') список файлов в директории
import locale
import openpyxl
import datetime
import calendar
import re

#получение из файла даты обучения
def get_diaposon_learning_date(path):
    book = openpyxl.load_workbook(path,data_only=True)
    sheet =  book.active
    cell = sheet.cell(row=1, column=1)
    start_str = cell.value.split('с ')[1].split(' по ')[0] #начало обучения
    end_str = cell.value.split('по ')[1] #конец обучения 
    return start_str,end_str

#получаем день недели str
def get_day_of_week(date_string):
    date_object = datetime.datetime.strptime(date_string, '%d.%m.%Y')
    locale.setlocale(locale.LC_ALL, 'ru_RU.UTF-8')
    day_of_week = date_object.strftime('%A')
    return day_of_week

#четная неделя или нет + там есть порядковый номер недели
def is_even_week(date_str, path):
    first_date, second_date = get_diaposon_learning_date(path)
    start_date = datetime.datetime.strptime(first_date, '%d.%m.%Y').date()
    end_date = datetime.datetime.strptime(second_date, '%d.%m.%Y').date()
    date_object = datetime.datetime.strptime(date_str, '%d.%m.%Y').date()
    if date_object < start_date or date_object > end_date:
        return "Мы не учимся брат"
    start_week_day = start_date.isoweekday()
    diff_days = (date_object - start_date).days
    week_number = (diff_days + start_week_day - 1) // 7 + 1 #какая учебная неделя
    is_even = week_number % 2 == 0
    return is_even



#получение ячейки группы в файле 
def find_group_in_file(group , path):
    book = openpyxl.load_workbook(path) # path
    sheet = book.active     #активный лист
    row = sheet[2] 
    
    for cell in row:
        result = (cell) 
        if cell.value == group:
            result = (cell) 
            return result
    return None

#Вывод в родительном падеже
def get_month_name_genitive(date):
    month_names_genitive = [
        "января",
        "февраля",
        "марта",
        "апреля",
        "мая",
        "июня",
        "июля",
        "августа",
        "сентября",
        "октября",
        "ноября",
        "декабря"
    ]

    date = datetime.datetime.strptime(date, '%d.%m.%Y')
    month_number = date.month - 1

    return f'{date.day} {month_names_genitive[month_number]}' 

#получение листа с расписанем на день расписания 
def get_raspisanie(cell, path, date): #ячейка, путь к файлу xml, дата
    locale.setlocale(locale.LC_ALL, 'ru_RU.UTF-8')
    book = openpyxl.load_workbook(path, read_only=True)
    sheet =  book.active
    base_row = cell.row
    base_column = cell.column
    uneven = is_even_week(date,path)
    day_of_the_week = get_day_of_week(date)
    #определение начального положения положения клетки(зависит от дня) 
    days_of_week = {
        "понедельник": 0,
        "вторник": 1,
        "среда": 2,
        "четверг": 3,
        "пятница": 4,
        "суббота": 5,
        "воскресенье": 6
    }
    #вывод расписания для этого дня
    formatted_date = get_month_name_genitive(date)
    date_for_list = f"Расписание на {formatted_date}:"
    value = days_of_week[day_of_the_week]
    if value == 6:
        return f"{date_for_list} \nПар нет, Чилим братья!"
    result_list = []
    result_list.append(date_for_list)
    shift_uneven = 0
    if uneven == False:
        pass
    else:
        shift_uneven += 1
    for i in range(0, 14, 2):
        row = [(sheet.cell(row=base_row + 14 * value + shift_uneven + i + 2, column=base_column + j).value or '-').replace("\n", " ")
            for j in range(4)]
        if row[0] == "-":
            result_list.append(f"{int(i/2 + 1)}) -")
        else:
            result_list.append(f"{int(i/2 + 1)}) {row[0]}, {row[1]}, {row[2]}, {row[3]}")
    result = '\n'.join(result_list)
    return result

#Возврат строки распиния.
def get_raspisanie_moment(date):
    group = "ИКБО-29-22"
    path = r"II\Oznokom praktika\BlockF\XLSX_files\IIT_1-kurs_22_23_vesna_27.04.2023.xlsx"
    cell = find_group_in_file(group, path)
    if cell is not None:
        return print(get_raspisanie(cell, path, date))
    else:
        return f"Группа {group} не найдена в файле"
    
get_raspisanie_moment('27.05.2023')